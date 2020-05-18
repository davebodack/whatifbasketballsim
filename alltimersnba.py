from openpyxl import load_workbook
from datetime import datetime
import operator
import requests
import random
import sys

class Team(object):

    def __init__(self):
        self.name = ""
        self.wins = 0
        self.seed = 0
        self.pointdiff = 0

    def __eq__(self, obj):
        return isinstance(obj, Team) and obj.name == self.name


#Runs one game between two teams
def run_game(hometeam, awayteam):

    hometeamfullsplit = hometeam.split(" ")
    awayteamfullsplit = awayteam.split(" ")
    requeststr = "https://www.whatifsports.com/NBA/default.asp?hSeason="
    requeststr += hometeamfullsplit[0]
    requeststr += "&hteam="
    for n in range(1, len(hometeamfullsplit)):
        if n + 1 < len(hometeamfullsplit):
            requeststr += (hometeamfullsplit[n] + "+")
        else:
            requeststr += (hometeamfullsplit[n] + "&")
    requeststr += "vSeason="
    requeststr += awayteamfullsplit[0]
    requeststr += "&vTeam="
    for n in range(1, len(awayteamfullsplit)):
        if n + 1 < len(awayteamfullsplit):
            requeststr += (awayteamfullsplit[n] + "+")
        else:
            requeststr += (awayteamfullsplit[n])
    requestget = requests.get(requeststr)
    gameid = requestget.text.split("GameID=")
    gameid = gameid[1].split("&", 1)
    gameid = gameid[0]

    requestboxscore = "https://www.whatifsports.com/NBA/pbp.asp?gameid="
    requestboxscore += gameid
    requestboxscore += "&qtr=4&teamfee=-1"
    boxscore = requests.get(requestboxscore)
    boxscore = boxscore.text

    nowrapindex = boxscore.rfind("<td nowrap>") + 11
    finalscore = boxscore[nowrapindex:].split('<')[0]
    print(gameid + ": " + awayteam + " " + finalscore + " " + hometeam)
    finalscore = finalscore.strip()
    return finalscore.split("-")


#Runs a full series of games between two teams
def run_series(hometeam, awayteam):

	for i in range(7):
		if (hometeam.wins < 4) and (awayteam.wins < 4):
			if (i < 2):
				scores = run_game(hometeam.name, awayteam.name)
			elif (i < 5):
				scores = run_game(awayteam.name, hometeam.name)
			else:
				scores = run_game(hometeam.name, awayteam.name)
			scores[0] = int(scores[0])
			scores[1] = int(scores[1])
			hometeam.pointdiff += scores[0] - scores[1]
			awayteam.pointdiff += scores[1] - scores[0]
			if ((i > 1) and (i < 5)):
				if (scores[0] > scores[1]):
					hometeam.wins += 1
				else:
					awayteam.wins += 1
			else:
				if (scores[1] > scores[0]):
					hometeam.wins += 1
				else:
					awayteam.wins += 1
	
		elif (hometeam.wins == 4):
			return hometeam.name
		else:
			return awayteam.name

	if (hometeam.wins == 4):
		return hometeam.name
	else:
		return awayteam.name


#All right, let's get this show on the road.
filepath = "./All-Timers NBA League.xlsx"
wb = load_workbook(filepath)
teamsheet = wb["Team List"]
bracket = wb["Year " + sys.argv[1] + " Bracket"]
finalssheet = wb["Yearly Finals Results"]
alltimesheet = wb["All-Time Results"]
nextyearbracket = wb["Year " + str(int(sys.argv[1]) + 1) + " Bracket"]

#Read the team names into the Team objects in the array 'teams'.
firstleagueteams = [Team() for i in range(16)]
secondleagueteams = [Team() for i in range(16)]
thirdleagueteams = [Team() for i in range(16)]
nextyearfirstleagueteams = [Team() for i in range(16)]
nextyearsecondleagueteams = [Team() for i in range(16)]
nextyearthirdleagueteams = [Team() for i in range(16)]

for i in range(16):
	firstleagueteams[i].name = teamsheet.cell(row = i + 1, column = 1).value
	firstleagueteams[i].seed = i // 2 + 1
	secondleagueteams[i].name = teamsheet.cell(row = i + 17, column = 1).value
	secondleagueteams[i].seed = i // 2 + 1
	thirdleagueteams[i].name = teamsheet.cell(row = i + 33, column = 1).value
	thirdleagueteams[i].seed = i // 2 + 1

print("---Welcome to the NBA Legends Playoffs!---")
print("")
print("--------------BRONZE LEAGUE---------------")
print("")

#These bad boys'll come in handy later.
thirdleagueroundtwoteams = [Team() for i in range(8)]
thirdleagueroundthreeteams = [Team() for i in range(4)]
thirdleagueroundfourteams = [Team() for i in range(2)]
thirdbracketroundonerows = [47, 47, 62, 62, 57, 57, 52, 52]
thirdbracketroundonecols = [3, 9, 9, 3, 3, 9, 9, 3]
thirdbracketroundtworows = [50, 50, 60, 60]
thirdbracketroundtwocols = [4, 8, 8, 4]
thirdbracketroundthreecols = [5, 7]


#Round One
print("         Conference Quarterfinals         ")
print("")

#There are twelve of these chunks of code between the three leagues, what each one does is run a series, send the winner to the next round, print what's happening on the terminal, update the Excel bracket, and figure out who goes where for the next year
for i in range(8):
	print(thirdleagueteams[i].name + " (" + str(thirdleagueteams[i].seed) + " seed)")
	print("                    VS.                   ")
	print(thirdleagueteams[15 - i].name + " (" + str(thirdleagueteams[15 - i].seed) + " seed)")
	print("")
	winner = run_series(thirdleagueteams[i], thirdleagueteams[15 - i]) 
	print("")
	if (thirdleagueteams[i].wins == 4):
		seriesscore = str(thirdleagueteams[i].wins) + "-" + str(thirdleagueteams[15 - i].wins)
		thirdleagueroundtwoteams[i] = thirdleagueteams[i]
	else:
		seriesscore = str(thirdleagueteams[15 - i].wins) + "-" + str(thirdleagueteams[i].wins)
		thirdleagueroundtwoteams[i] = thirdleagueteams[15 - i]
	print("WINNER: " + winner + " " + seriesscore) 
	print("")
	print("")
	bracket.cell(row = thirdbracketroundonerows[i], column = thirdbracketroundonecols[i]).value = winner
	if (((i > 0) and (i < 3)) or ((i > 4) and (i < 7))):
		bracket.cell(row = thirdbracketroundonerows[i], column = thirdbracketroundonecols[i] + 1).value = seriesscore
	else:
		bracket.cell(row = thirdbracketroundonerows[i], column = thirdbracketroundonecols[i] - 1).value = seriesscore

thirdleagueteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(8, 16):
	nextyearthirdleagueteams[i] = thirdleagueteams[i]


#Round Two
print("----------Conference Semifinals-----------")  
print("")

for i in range(8):
	thirdleagueroundtwoteams[i].wins = 0
	thirdleagueroundtwoteams[i].pointdiff = 0

for i in range(4):
	print(thirdleagueroundtwoteams[i].name + " (" + str(thirdleagueroundtwoteams[i].seed) + " seed)")
	print("                    VS.                   ")
	print(thirdleagueroundtwoteams[7 - i].name + " (" + str(thirdleagueroundtwoteams[7 - i].seed) + " seed)")
	print("")
	winner = run_series(thirdleagueroundtwoteams[i], thirdleagueroundtwoteams[7 - i]) 
	print("")
	if (thirdleagueroundtwoteams[i].wins == 4):
		seriesscore = str(thirdleagueroundtwoteams[i].wins) + "-" + str(thirdleagueroundtwoteams[7 - i].wins)
		thirdleagueroundthreeteams[i] = thirdleagueroundtwoteams[i]
	else:
		seriesscore = str(thirdleagueroundtwoteams[7 - i].wins) + "-" + str(thirdleagueroundtwoteams[i].wins)
		thirdleagueroundthreeteams[i] = thirdleagueroundtwoteams[7 - i]
	print("WINNER: " + winner + " " + seriesscore) 
	print("")
	print("")
	bracket.cell(row = thirdbracketroundtworows[i], column = thirdbracketroundtwocols[i]).value = winner
	if ((i > 0) and (i < 3)):
		bracket.cell(row = thirdbracketroundtworows[i], column = thirdbracketroundtwocols[i] + 1).value = seriesscore
	else:
		bracket.cell(row = thirdbracketroundtworows[i], column = thirdbracketroundtwocols[i] - 1).value = seriesscore

thirdleagueroundtwoteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(4, 8):
	nextyearthirdleagueteams[i] = thirdleagueroundtwoteams[i]


#Round Three
print("----------Conference Finals-----------")  
print("")

for i in range(4):
	thirdleagueroundthreeteams[i].wins = 0
	thirdleagueroundthreeteams[i].pointdiff = 0

for i in range(2):
	print(thirdleagueroundthreeteams[i].name + " (" + str(thirdleagueroundthreeteams[i].seed) + " seed)")
	print("                    VS.                   ")
	print(thirdleagueroundthreeteams[3 - i].name + " (" + str(thirdleagueroundthreeteams[3 - i].seed) + " seed)")
	print("")
	winner = run_series(thirdleagueroundthreeteams[i], thirdleagueroundthreeteams[3 - i]) 
	print("")
	if (thirdleagueroundthreeteams[i].wins == 4):
		seriesscore = str(thirdleagueroundthreeteams[i].wins) + "-" + str(thirdleagueroundthreeteams[3 - i].wins)
		thirdleagueroundfourteams[i] = thirdleagueroundthreeteams[i]
	else:
		seriesscore = str(thirdleagueroundthreeteams[3 - i].wins) + "-" + str(thirdleagueroundthreeteams[i].wins)
		thirdleagueroundfourteams[i] = thirdleagueroundthreeteams[3 - i]
	print("WINNER: " + winner + " " + seriesscore) 
	print("")
	print("")
	bracket.cell(row = 55, column = thirdbracketroundthreecols[i]).value = winner
	if (i == 1):
		bracket.cell(row = 55, column = thirdbracketroundthreecols[i] + 1).value = seriesscore
	else:
		bracket.cell(row = 55, column = thirdbracketroundthreecols[i] - 1).value = seriesscore

thirdleagueroundthreeteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(2, 4):
	nextyearthirdleagueteams[i] = thirdleagueroundthreeteams[i]


#Round Four
print("---------BRONZE LEAGUE NBA FINALS---------")  
print("")

for i in range(2):
	thirdleagueroundfourteams[i].wins = 0

print(thirdleagueroundfourteams[0].name + " (" + str(thirdleagueroundfourteams[0].seed) + " seed)")
print("                    VS.                   ")
print(thirdleagueroundfourteams[1].name + " (" + str(thirdleagueroundfourteams[1].seed) + " seed)")
print("")
if (thirdleagueroundfourteams[0].seed > thirdleagueroundfourteams[1].seed):
	winner = run_series(thirdleagueroundfourteams[0], thirdleagueroundfourteams[1]) 
elif (thirdleagueroundfourteams[1].seed > thirdleagueroundfourteams[0].seed):
	winner = run_series(thirdleagueroundfourteams[1], thirdleagueroundfourteams[0])
else:
	if (thirdleagueroundfourteams[0].pointdiff > thirdleagueroundfourteams[1].pointdiff):
		winner = run_series(thirdleagueroundfourteams[0], thirdleagueroundfourteams[1])
	else:
		winner = run_series(thirdleagueroundfourteams[1], thirdleagueroundfourteams[0])

print("")
if (thirdleagueroundfourteams[0].wins == 4):
	seriesscore = str(thirdleagueroundfourteams[0].wins) + "-" + str(thirdleagueroundfourteams[1].wins)
else:
	seriesscore = str(thirdleagueroundfourteams[1].wins) + "-" + str(thirdleagueroundfourteams[0].wins)

print("WINNER: " + winner + " " + seriesscore) 
print("")
print("")
bracket.cell(row = 58, column = 6).value = winner
bracket.cell(row = 59, column = 6).value = seriesscore

thirdleagueroundfourteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(2):
	nextyearsecondleagueteams[i + 14] = thirdleagueroundfourteams[i]

finalssheet.cell(row = int(sys.argv[1]) + 2, column = 8).value = nextyearsecondleagueteams[14].name
finalssheet.cell(row = int(sys.argv[1]) + 2, column = 9).value = seriesscore
finalssheet.cell(row = int(sys.argv[1]) + 2, column = 10).value = nextyearsecondleagueteams[15].name


print("")
print("--------------SILVER LEAGUE---------------")
print("")

secondleagueroundtwoteams = [Team() for i in range(8)]
secondleagueroundthreeteams = [Team() for i in range(4)]
secondleagueroundfourteams = [Team() for i in range(2)]
secondbracketroundonerows = [25, 25, 40, 40, 35, 35, 30, 30]
secondbracketroundonecols = [3, 9, 9, 3, 3, 9, 9, 3]
secondbracketroundtworows = [28, 28, 38, 38]
secondbracketroundtwocols = [4, 8, 8, 4]
secondbracketroundthreecols = [5, 7]

#Round One
print("         Conference Quarterfinals         ")
print("")

for i in range(8):
	print(secondleagueteams[i].name + " (" + str(secondleagueteams[i].seed) + " seed)")
	print("                    VS.                   ")
	print(secondleagueteams[15 - i].name + " (" + str(secondleagueteams[15 - i].seed) + " seed)")
	print("")
	winner = run_series(secondleagueteams[i], secondleagueteams[15 - i]) 
	print("")
	if (secondleagueteams[i].wins == 4):
		seriesscore = str(secondleagueteams[i].wins) + "-" + str(secondleagueteams[15 - i].wins)
		secondleagueroundtwoteams[i] = secondleagueteams[i]
	else:
		seriesscore = str(secondleagueteams[15 - i].wins) + "-" + str(secondleagueteams[i].wins)
		secondleagueroundtwoteams[i] = secondleagueteams[15 - i]
	print("WINNER: " + winner + " " + seriesscore) 
	print("")
	print("")
	bracket.cell(row = secondbracketroundonerows[i], column = secondbracketroundonecols[i]).value = winner
	if (((i > 0) and (i < 3)) or ((i > 4) and (i < 7))):
		bracket.cell(row = secondbracketroundonerows[i], column = secondbracketroundonecols[i] + 1).value = seriesscore
	else:
		bracket.cell(row = secondbracketroundonerows[i], column = secondbracketroundonecols[i] - 1).value = seriesscore

secondleagueteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(2):
	nextyearthirdleagueteams[i] = secondleagueteams[i + 14]
for i in range(8, 14):
	nextyearsecondleagueteams[i] = secondleagueteams[i]


#Round Two
print("----------Conference Semifinals-----------")  
print("")

for i in range(8):
	secondleagueroundtwoteams[i].wins = 0
	secondleagueroundtwoteams[i].pointdiff = 0

for i in range(4):
	print(secondleagueroundtwoteams[i].name + " (" + str(secondleagueroundtwoteams[i].seed) + " seed)")
	print("                    VS.                   ")
	print(secondleagueroundtwoteams[7 - i].name + " (" + str(secondleagueroundtwoteams[7 - i].seed) + " seed)")
	print("")
	winner = run_series(secondleagueroundtwoteams[i], secondleagueroundtwoteams[7 - i]) 
	print("")
	if (secondleagueroundtwoteams[i].wins == 4):
		seriesscore = str(secondleagueroundtwoteams[i].wins) + "-" + str(secondleagueroundtwoteams[7 - i].wins)
		secondleagueroundthreeteams[i] = secondleagueroundtwoteams[i]
	else:
		seriesscore = str(secondleagueroundtwoteams[7 - i].wins) + "-" + str(secondleagueroundtwoteams[i].wins)
		secondleagueroundthreeteams[i] = secondleagueroundtwoteams[7 - i]
	print("WINNER: " + winner + " " + seriesscore) 
	print("")
	print("")
	bracket.cell(row = secondbracketroundtworows[i], column = secondbracketroundtwocols[i]).value = winner
	if ((i > 0) and (i < 3)):
		bracket.cell(row = secondbracketroundtworows[i], column = secondbracketroundtwocols[i] + 1).value = seriesscore
	else:
		bracket.cell(row = secondbracketroundtworows[i], column = secondbracketroundtwocols[i] - 1).value = seriesscore

secondleagueroundtwoteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(4, 8):
	nextyearsecondleagueteams[i] = secondleagueroundtwoteams[i]


#Round Three
print("----------Conference Finals-----------")  
print("")

for i in range(4):
	secondleagueroundthreeteams[i].wins = 0
	secondleagueroundthreeteams[i].pointdiff = 0

for i in range(2):
	print(secondleagueroundthreeteams[i].name + " (" + str(secondleagueroundthreeteams[i].seed) + " seed)")
	print("                    VS.                   ")
	print(secondleagueroundthreeteams[3 - i].name + " (" + str(secondleagueroundthreeteams[3 - i].seed) + " seed)")
	print("")
	winner = run_series(secondleagueroundthreeteams[i], secondleagueroundthreeteams[3 - i]) 
	print("")
	if (secondleagueroundthreeteams[i].wins == 4):
		seriesscore = str(secondleagueroundthreeteams[i].wins) + "-" + str(secondleagueroundthreeteams[3 - i].wins)
		secondleagueroundfourteams[i] = secondleagueroundthreeteams[i]
	else:
		seriesscore = str(secondleagueroundthreeteams[3 - i].wins) + "-" + str(secondleagueroundthreeteams[i].wins)
		secondleagueroundfourteams[i] = secondleagueroundthreeteams[3 - i]
	print("WINNER: " + winner + " " + seriesscore) 
	print("")
	print("")
	bracket.cell(row = 33, column = secondbracketroundthreecols[i]).value = winner
	if (i == 1):
		bracket.cell(row = 33, column = secondbracketroundthreecols[i] + 1).value = seriesscore
	else:
		bracket.cell(row = 33, column = secondbracketroundthreecols[i] - 1).value = seriesscore

secondleagueroundthreeteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(2, 4):
	nextyearsecondleagueteams[i] = secondleagueroundthreeteams[i]


#Round Four
print("---------SILVER LEAGUE NBA FINALS---------")  
print("")

for i in range(2):
	secondleagueroundfourteams[i].wins = 0

print(secondleagueroundfourteams[0].name + " (" + str(secondleagueroundfourteams[0].seed) + " seed)")
print("                    VS.                   ")
print(secondleagueroundfourteams[1].name + " (" + str(secondleagueroundfourteams[1].seed) + " seed)")
print("")
if (secondleagueroundfourteams[0].seed > secondleagueroundfourteams[1].seed):
	winner = run_series(secondleagueroundfourteams[0], secondleagueroundfourteams[1]) 
elif (secondleagueroundfourteams[1].seed > secondleagueroundfourteams[0].seed):
	winner = run_series(secondleagueroundfourteams[1], secondleagueroundfourteams[0])
else:
	if (secondleagueroundfourteams[0].pointdiff > secondleagueroundfourteams[1].pointdiff):
		winner = run_series(secondleagueroundfourteams[0], secondleagueroundfourteams[1])
	else:
		winner = run_series(secondleagueroundfourteams[1], secondleagueroundfourteams[0])

print("")
if (secondleagueroundfourteams[0].wins == 4):
	seriesscore = str(secondleagueroundfourteams[0].wins) + "-" + str(secondleagueroundfourteams[1].wins)
else:
	seriesscore = str(secondleagueroundfourteams[1].wins) + "-" + str(secondleagueroundfourteams[0].wins)

print("WINNER: " + winner + " " + seriesscore) 
print("")
print("")
bracket.cell(row = 36, column = 6).value = winner
bracket.cell(row = 37, column = 6).value = seriesscore

secondleagueroundfourteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(2):
	nextyearfirstleagueteams[i + 14] = secondleagueroundfourteams[i]

finalssheet.cell(row = int(sys.argv[1]) + 2, column = 5).value = nextyearfirstleagueteams[14].name
finalssheet.cell(row = int(sys.argv[1]) + 2, column = 6).value = seriesscore
finalssheet.cell(row = int(sys.argv[1]) + 2, column = 7).value = nextyearfirstleagueteams[15].name


print("")
print("--------------GOLD LEAGUE---------------")
print("")

firstleagueroundtwoteams = [Team() for i in range(8)]
firstleagueroundthreeteams = [Team() for i in range(4)]
firstleagueroundfourteams = [Team() for i in range(2)]
firstbracketroundonerows = [3, 3, 18, 18, 13, 13, 8, 8]
firstbracketroundonecols = [3, 9, 9, 3, 3, 9, 9, 3]
firstbracketroundtworows = [6, 6, 16, 16]
firstbracketroundtwocols = [4, 8, 8, 4]
firstbracketroundthreecols = [5, 7]

#Round One
print("         Conference Quarterfinals         ")
print("")

for i in range(8):
	print(firstleagueteams[i].name + " (" + str(firstleagueteams[i].seed) + " seed)")
	print("                    VS.                   ")
	print(firstleagueteams[15 - i].name + " (" + str(firstleagueteams[15 - i].seed) + " seed)")
	print("")
	winner = run_series(firstleagueteams[i], firstleagueteams[15 - i]) 
	print("")
	if (firstleagueteams[i].wins == 4):
		seriesscore = str(firstleagueteams[i].wins) + "-" + str(firstleagueteams[15 - i].wins)
		firstleagueroundtwoteams[i] = firstleagueteams[i]
	else:
		seriesscore = str(firstleagueteams[15 - i].wins) + "-" + str(firstleagueteams[i].wins)
		firstleagueroundtwoteams[i] = firstleagueteams[15 - i]
	print("WINNER: " + winner + " " + seriesscore) 
	print("")
	print("")
	bracket.cell(row = firstbracketroundonerows[i], column = firstbracketroundonecols[i]).value = winner
	if (((i > 0) and (i < 3)) or ((i > 4) and (i < 7))):
		bracket.cell(row = firstbracketroundonerows[i], column = firstbracketroundonecols[i] + 1).value = seriesscore
	else:
		bracket.cell(row = firstbracketroundonerows[i], column = firstbracketroundonecols[i] - 1).value = seriesscore

firstleagueteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(2):
	nextyearsecondleagueteams[i] = firstleagueteams[i + 14]
for i in range(8, 14):
	nextyearfirstleagueteams[i] = firstleagueteams[i]


#Round Two
print("----------Conference Semifinals-----------")  
print("")

for i in range(8):
	firstleagueroundtwoteams[i].wins = 0
	firstleagueroundtwoteams[i].pointdiff = 0

for i in range(4):
	print(firstleagueroundtwoteams[i].name + " (" + str(firstleagueroundtwoteams[i].seed) + " seed)")
	print("                    VS.                   ")
	print(firstleagueroundtwoteams[7 - i].name + " (" + str(firstleagueroundtwoteams[7 - i].seed) + " seed)")
	print("")
	winner = run_series(firstleagueroundtwoteams[i], firstleagueroundtwoteams[7 - i]) 
	print("")
	if (firstleagueroundtwoteams[i].wins == 4):
		seriesscore = str(firstleagueroundtwoteams[i].wins) + "-" + str(firstleagueroundtwoteams[7 - i].wins)
		firstleagueroundthreeteams[i] = firstleagueroundtwoteams[i]
	else:
		seriesscore = str(firstleagueroundtwoteams[7 - i].wins) + "-" + str(firstleagueroundtwoteams[i].wins)
		firstleagueroundthreeteams[i] = firstleagueroundtwoteams[7 - i]
	print("WINNER: " + winner + " " + seriesscore) 
	print("")
	print("")
	bracket.cell(row = firstbracketroundtworows[i], column = firstbracketroundtwocols[i]).value = winner
	if ((i > 0) and (i < 3)):
		bracket.cell(row = firstbracketroundtworows[i], column = firstbracketroundtwocols[i] + 1).value = seriesscore
	else:
		bracket.cell(row = firstbracketroundtworows[i], column = firstbracketroundtwocols[i] - 1).value = seriesscore

firstleagueroundtwoteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(4, 8):
	nextyearfirstleagueteams[i] = firstleagueroundtwoteams[i]


#Round Three
print("----------Conference Finals-----------")  
print("")

for i in range(4):
	firstleagueroundthreeteams[i].wins = 0
	firstleagueroundthreeteams[i].pointdiff = 0

for i in range(2):
	print(firstleagueroundthreeteams[i].name + " (" + str(firstleagueroundthreeteams[i].seed) + " seed)")
	print("                    VS.                   ")
	print(firstleagueroundthreeteams[3 - i].name + " (" + str(firstleagueroundthreeteams[3 - i].seed) + " seed)")
	print("")
	winner = run_series(firstleagueroundthreeteams[i], firstleagueroundthreeteams[3 - i]) 
	print("")
	if (firstleagueroundthreeteams[i].wins == 4):
		seriesscore = str(firstleagueroundthreeteams[i].wins) + "-" + str(firstleagueroundthreeteams[3 - i].wins)
		firstleagueroundfourteams[i] = firstleagueroundthreeteams[i]
	else:
		seriesscore = str(firstleagueroundthreeteams[3 - i].wins) + "-" + str(firstleagueroundthreeteams[i].wins)
		firstleagueroundfourteams[i] = firstleagueroundthreeteams[3 - i]
	print("WINNER: " + winner + " " + seriesscore) 
	print("")
	print("")
	bracket.cell(row = 11, column = firstbracketroundthreecols[i]).value = winner
	if (i == 1):
		bracket.cell(row = 11, column = firstbracketroundthreecols[i] + 1).value = seriesscore
	else:
		bracket.cell(row = 11, column = firstbracketroundthreecols[i] - 1).value = seriesscore

firstleagueroundthreeteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(2, 4):
	nextyearfirstleagueteams[i] = firstleagueroundthreeteams[i]


#Round Four
print("---------GOLD LEAGUE NBA FINALS---------")  
print("")

for i in range(2):
	firstleagueroundfourteams[i].wins = 0

print(firstleagueroundfourteams[0].name + " (" + str(firstleagueroundfourteams[0].seed) + " seed)")
print("                    VS.                   ")
print(firstleagueroundfourteams[1].name + " (" + str(firstleagueroundfourteams[1].seed) + " seed)")
print("")
if (firstleagueroundfourteams[0].seed > firstleagueroundfourteams[1].seed):
	winner = run_series(firstleagueroundfourteams[0], firstleagueroundfourteams[1]) 
elif (firstleagueroundfourteams[1].seed > firstleagueroundfourteams[0].seed):
	winner = run_series(firstleagueroundfourteams[1], firstleagueroundfourteams[0])
else:
	if (firstleagueroundfourteams[0].pointdiff > firstleagueroundfourteams[1].pointdiff):
		winner = run_series(firstleagueroundfourteams[0], firstleagueroundfourteams[1])
	else:
		winner = run_series(firstleagueroundfourteams[1], firstleagueroundfourteams[0])

print("")
if (firstleagueroundfourteams[0].wins == 4):
	seriesscore = str(firstleagueroundfourteams[0].wins) + "-" + str(firstleagueroundfourteams[1].wins)
else:
	seriesscore = str(firstleagueroundfourteams[1].wins) + "-" + str(firstleagueroundfourteams[0].wins)

print("WINNER: " + winner + " " + seriesscore) 
print("")
print("")
bracket.cell(row = 14, column = 6).value = winner
bracket.cell(row = 15, column = 6).value = seriesscore

firstleagueroundfourteams.sort(reverse = True, key = lambda Team: (Team.wins, Team.pointdiff))
for i in range(2):
	nextyearfirstleagueteams[i] = firstleagueroundfourteams[i]

finalssheet.cell(row = int(sys.argv[1]) + 2, column = 2).value = nextyearfirstleagueteams[0].name
finalssheet.cell(row = int(sys.argv[1]) + 2, column = 3).value = seriesscore
finalssheet.cell(row = int(sys.argv[1]) + 2, column = 4).value = nextyearfirstleagueteams[1].name


#Update All-Time Results sheet
for i in range(2, 50):

	#Update number of years in each league
	for j in range(16):
		if alltimesheet.cell(row = i, column = 2).value == firstleagueteams[j].name:
			alltimesheet.cell(row = i, column = 4).value = str(int(alltimesheet.cell(row = i, column = 4).value) + 1)
		if alltimesheet.cell(row = i, column = 2).value == secondleagueteams[j].name:
			alltimesheet.cell(row = i, column = 7).value = str(int(alltimesheet.cell(row = i, column = 7).value) + 1)
		if alltimesheet.cell(row = i, column = 2).value == thirdleagueteams[j].name:
			alltimesheet.cell(row = i, column = 10).value = str(int(alltimesheet.cell(row = i, column = 10).value) + 1)
	
	#Update Gold League Finals appearances and champs
	if alltimesheet.cell(row = i, column = 2).value == nextyearfirstleagueteams[0].name:
		alltimesheet.cell(row = i, column = 5).value = str(int(alltimesheet.cell(row = i, column = 5).value) + 1)
		alltimesheet.cell(row = i, column = 6).value = str(int(alltimesheet.cell(row = i, column = 6).value) + 1)
	if alltimesheet.cell(row = i, column = 2).value == nextyearfirstleagueteams[1].name:
		alltimesheet.cell(row = i, column = 6).value = str(int(alltimesheet.cell(row = i, column = 6).value) + 1)

	#Update Silver League Finals appearances and champs
	if alltimesheet.cell(row = i, column = 2).value == nextyearfirstleagueteams[14].name:
		alltimesheet.cell(row = i, column = 8).value = str(int(alltimesheet.cell(row = i, column = 8).value) + 1)
		alltimesheet.cell(row = i, column = 9).value = str(int(alltimesheet.cell(row = i, column = 9).value) + 1)
	if alltimesheet.cell(row = i, column = 2).value == nextyearfirstleagueteams[15].name:
		alltimesheet.cell(row = i, column = 9).value = str(int(alltimesheet.cell(row = i, column = 9).value) + 1)

	#Update Bronze League Finals appearances and champs
	if alltimesheet.cell(row = i, column = 2).value == nextyearsecondleagueteams[14].name:
		alltimesheet.cell(row = i, column = 11).value = str(int(alltimesheet.cell(row = i, column = 11).value) + 1)
		alltimesheet.cell(row = i, column = 12).value = str(int(alltimesheet.cell(row = i, column = 12).value) + 1)
	if alltimesheet.cell(row = i, column = 2).value == nextyearsecondleagueteams[15].name:
		alltimesheet.cell(row = i, column = 12).value = str(int(alltimesheet.cell(row = i, column = 12).value) + 1)

	#Update promoted and demoted teams
	if (alltimesheet.cell(row = i, column = 2).value == nextyearsecondleagueteams[14].name) or (alltimesheet.cell(row = i, column = 2).value == nextyearsecondleagueteams[15].name):
		alltimesheet.cell(row = i, column = 13).value = str(int(alltimesheet.cell(row = i, column = 13).value) + 1)
	if (alltimesheet.cell(row = i, column = 2).value == nextyearfirstleagueteams[14].name) or (alltimesheet.cell(row = i, column = 2).value == nextyearfirstleagueteams[15].name):
		alltimesheet.cell(row = i, column = 13).value = str(int(alltimesheet.cell(row = i, column = 13).value) + 1)
	if (alltimesheet.cell(row = i, column = 2).value == nextyearsecondleagueteams[0].name) or (alltimesheet.cell(row = i, column = 2).value == nextyearsecondleagueteams[1].name):
		alltimesheet.cell(row = i, column = 14).value = str(int(alltimesheet.cell(row = i, column = 14).value) + 1)
	if (alltimesheet.cell(row = i, column = 2).value == nextyearthirdleagueteams[0].name) or (alltimesheet.cell(row = i, column = 2).value == nextyearthirdleagueteams[1].name):
		alltimesheet.cell(row = i, column = 14).value = str(int(alltimesheet.cell(row = i, column = 14).value) + 1)


#Set up next year's bracket and team list(You have to make next year's bracket before you run this year's. Sorry.)
firstrdcols = [2, 10, 10, 2, 2, 10, 10, 2, 2, 10, 10, 2, 2, 10, 10, 2]
firstrdrows = [2, 2, 17, 17, 12, 12, 7, 7, 9, 9, 14, 14, 19, 19, 4, 4]
for i in range(16):
	nextyearbracket.cell(row = i + 3, column = 12).value = nextyearfirstleagueteams[i].name
	nextyearbracket.cell(row = firstrdrows[i], column = firstrdcols[i]).value = nextyearfirstleagueteams[i].name
	teamsheet.cell(row = i + 1, column = 1).value = nextyearfirstleagueteams[i].name
	nextyearbracket.cell(row = i + 25, column = 12).value = nextyearsecondleagueteams[i].name 
	nextyearbracket.cell(row = firstrdrows[i] + 22, column = firstrdcols[i]).value = nextyearsecondleagueteams[i].name
	teamsheet.cell(row = i + 17, column = 1).value = nextyearsecondleagueteams[i].name
	nextyearbracket.cell(row = i + 47, column = 12).value = nextyearthirdleagueteams[i].name
	nextyearbracket.cell(row = firstrdrows[i] + 44, column = firstrdcols[i]).value = nextyearthirdleagueteams[i].name
	teamsheet.cell(row = i + 33, column = 1).value = nextyearthirdleagueteams[i].name


wb.save(filepath)